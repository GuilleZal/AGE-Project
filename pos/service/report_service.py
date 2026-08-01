"""Report service — sales summaries, profit analysis, and CSV export.

Performance target: <3s for 10k sales (verified via date indexes on
``sales.created_at`` and ``sale_items`` join columns).
"""

import csv
import os
import sqlite3

from pos.repository.product_repo import ProductRepo
from pos.repository.sale_repo import SaleRepo


class ReportService:
    """Aggregation logic for sales and profit reports.

    Wraps ``SaleRepo`` for data access. All summary methods return dicts
    with type-hinted keys for easy consumption by controllers and views.
    """

    def __init__(self, db: sqlite3.Connection) -> None:
        self._db = db
        self._sale_repo = SaleRepo(db)
        self._product_repo = ProductRepo(db)

    # -------------------------------------------------------- sales summary

    def sales_summary(
        self, start_date: str, end_date: str
    ) -> dict[str, int | float]:
        """Return total revenue, sale count and average ticket.

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            Dict with keys ``total`` (int), ``count`` (int), ``avg_ticket`` (float).
            Division-by-zero safe: ``avg_ticket`` is 0.0 when ``count`` is 0.
        """
        row = self._db.execute(
            """SELECT COALESCE(SUM(total), 0) AS total,
                       COUNT(*)              AS count
               FROM sales
               WHERE created_at >= ? AND created_at <= ?""",
            (start_date, end_date),
        ).fetchone()

        total = row["total"] or 0
        count = row["count"] or 0
        avg_ticket = (total / count) if count > 0 else 0.0
        return {"total": total, "count": count, "avg_ticket": avg_ticket}

    # ------------------------------------------------------- profit summary

    def profit_summary(
        self, start_date: str, end_date: str
    ) -> dict[str, int | float]:
        """Return revenue, cost, profit and margin percentage.

        Revenue = sum(sale_items.subtotal)
        Cost    = sum(sale_items.quantity × product.cost_price)
        Profit  = revenue − cost
        Margin% = (profit / revenue) × 100 (0 when revenue ≤ 0)

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            Dict with keys ``revenue``, ``cost``, ``profit``, ``margin_pct``.
        """
        row = self._db.execute(
            """SELECT COALESCE(SUM(si.subtotal), 0)       AS revenue,
                      COALESCE(SUM(si.quantity * p.cost_price), 0) AS cost
               FROM sale_items si
               JOIN products p ON p.id = si.product_id
               JOIN sales    s ON s.id = si.sale_id
                WHERE s.created_at >= ? AND s.created_at <= ?""",
            (start_date, end_date),
        ).fetchone()

        revenue = row["revenue"] or 0
        cost = row["cost"] or 0
        profit = revenue - cost
        margin_pct = (profit / revenue * 100) if revenue > 0 else 0.0
        return {
            "revenue": revenue,
            "cost": cost,
            "profit": profit,
            "margin_pct": margin_pct,
        }

    # --------------------------------------------------------- top products

    def top_products(
        self, start_date: str, end_date: str, limit: int = 10
    ) -> list[dict]:
        """Return the top *N* products sold in the period.

        Delegates to ``SaleRepo.top_products`` which provides
        ``product_id``, ``name``, ``barcode``, ``total_quantity``,
        ``total_amount``.

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).
            limit:      Max number of products to return (default 10).

        Returns:
            List of product summary dicts sorted by total_quantity DESC.
        """
        return self._sale_repo.top_products(start_date, end_date, limit)

    # ------------------------------------------------------- low stock ----

    def low_stock_products(self) -> list[dict]:
        """Return all products whose current stock is at or below their
        ``low_stock_threshold``.

        Delegates to ``ProductRepo.low_stock_products``.

        Returns:
            List of dicts with product fields, ordered by stock ascending.
        """
        return self._product_repo.low_stock_products()

    # ---------------------------------------------- payment methods summary

    _PAYMENT_METHOD_LABELS: dict[str, str] = {
        "cash": "Efectivo",
        "debit_card": "Tarjeta de Débito",
        "credit_card": "Tarjeta de Crédito",
        "transfer": "Transferencia",
        "qr": "Qr",
    }

    def payment_methods_summary(
        self, start_date: str, end_date: str
    ) -> list[dict]:
        """Return revenue breakdown by payment method for the given period.

        Each entry includes the absolute total, operations count and its percentage
        of the grand total.
        
        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            List of dicts with keys ``payment_method`` (str),
            ``sale_count`` (int), ``total_amount`` (int), ``percentage`` (float).
            Sorted by ``total_amount`` DESC.
        """
        raw = self._sale_repo.total_by_payment_method(start_date, end_date)
        
        grand_total = sum(item["total_amount"] for item in raw)

        result: list[dict] = []
        for item in raw:
            method = item["payment_method"]
            total_amount = item["total_amount"]
            sale_count = item.get("sale_count", 0)
            percentage = (
                round((total_amount / grand_total) * 100, 1)
                if grand_total > 0
                else 0.0
            )
            result.append(
                {
                    "payment_method": self._PAYMENT_METHOD_LABELS.get(
                        method, method
                    ),
                    "sale_count": sale_count,
                    "total_amount": total_amount,
                    "percentage": percentage,
                }
            )

        result.sort(key=lambda x: x["total_amount"], reverse=True)
        return result

    # ---------------------------------------------------- sales by category

    def sales_by_category(
        self, start_date: str, end_date: str
    ) -> list[dict]:
        """Return sales aggregate by product category for the given period.

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            List of dicts with keys: ``category_name``, ``qty_kg``, ``qty_unit``,
            ``total_amount``.
        """
        rows = self._db.execute(
            """SELECT COALESCE(c.name, 'Sin categoría') AS category_name,
                       SUM(CASE WHEN p.unit_type = 'Kg' THEN si.quantity ELSE 0 END) AS qty_kg,
                       SUM(CASE WHEN p.unit_type = 'Unidad' THEN si.quantity ELSE 0 END) AS qty_unit,
                       SUM(si.subtotal) AS total_amount
                FROM sale_items si
                JOIN products p ON p.id = si.product_id
                LEFT JOIN categories c ON c.id = p.category_id
                JOIN sales s ON s.id = si.sale_id
                WHERE s.created_at >= ? AND s.created_at <= ?
                GROUP BY p.category_id
                ORDER BY total_amount DESC""",
            (start_date, end_date),
        ).fetchall()
        return [dict(r) for r in rows]

    # ------------------------------------------------------ returns history

    def returns_history(
        self, start_date: str, end_date: str
    ) -> list[dict]:
        """Return returns history for the given period with product details.

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            List of dicts with keys: ``created_at``, ``product_name``,
            ``quantity``, ``refund_amount``, ``reason``.
        """
        rows = self._db.execute(
            """SELECT r.created_at,
                       p.name AS product_name,
                       r.quantity,
                       r.refund_amount,
                       r.reason
                FROM returns r
                JOIN products p ON p.id = r.product_id
                WHERE r.created_at >= ? AND r.created_at <= ?
                ORDER BY r.created_at DESC""",
            (start_date, end_date),
        ).fetchall()
        return [dict(r) for r in rows]

    # ----------------------------------------------- purchases summary ----

    def purchases_summary(
        self, start_date: str, end_date: str
    ) -> dict[str, int]:
        """Return total purchase spend for the given period.

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            Dict with key ``total`` (int).
        """
        row = self._db.execute(
            """SELECT COALESCE(SUM(total), 0) AS total
               FROM purchases
               WHERE purchase_date >= ? AND purchase_date <= ?""",
            (start_date, end_date),
        ).fetchone()
        return {"total": row["total"] or 0}

    # ------------------------------------------------- expenses summary ---

    def expenses_summary(
        self, start_date: str, end_date: str
    ) -> dict[str, int]:
        """Return expense components for the given period.

        Components (from cash_movements table):
            * purchases — total of 'supplier_payment' movements (Pago a Proveedor).
            * shrinkage — kept for compatibility (always 0, no source yet).
            * operating_expenses — total of 'expense' movements (Gastos).

        Args:
            start_date: ISO-like datetime string (inclusive).
            end_date:   ISO-like datetime string (inclusive).

        Returns:
            Dict with keys ``purchases``, ``shrinkage``,
            ``operating_expenses`` (all int).
        """
        # Compras a Proveedores = supplier_payment from cash_movements
        purchases = self._db.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM cash_movements
               WHERE type = 'supplier_payment'
                 AND created_at >= ? AND created_at <= ?""",
            (start_date, end_date),
        ).fetchone()

        # Gastos Operativos = expense from cash_movements
        operating = self._db.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM cash_movements
               WHERE type = 'expense'
                 AND created_at >= ? AND created_at <= ?""",
            (start_date, end_date),
        ).fetchone()

        # Pérdidas (shrinkage) = returns with reasons other than "Producto en buenas condiciones"
        shrinkage = self._db.execute(
            """SELECT COALESCE(SUM(refund_amount), 0) AS total
               FROM returns
               WHERE (reason IS NULL OR reason != 'Producto en buenas condiciones')
                 AND created_at >= ? AND created_at <= ?""",
            (start_date, end_date),
        ).fetchone()

        return {
            "purchases": purchases["total"] or 0,
            "shrinkage": shrinkage["total"] or 0,
            "operating_expenses": operating["total"] or 0,
        }

    # -------------------------------------------------------------- CSV ----

    @staticmethod
    def export_csv(data: list[dict], filepath: str, start_date: str = "", end_date: str = "", title: str = "") -> str:
        """Write *data* (list of dicts) to a semicolon-delimited CSV with BOM.

        The UTF-8 BOM ensures Excel (Spanish locale) opens the file correctly.
        Semicolons are used instead of commas per Argentine Excel convention.

        Args:
            data:     List of homogeneous dicts (all same keys).
            filepath: Destination path for the CSV file.
            start_date: Start date of the report range.
            end_date:   End date of the report range.
            title:      Optional title for the report.

        Returns:
            The *filepath* on success.

        Raises:
            OSError: If the file cannot be written.
        """
        import os
        import csv
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            if title:
                f.write(f"{title}\n")
            if start_date and end_date:
                f.write(f"Reporte desde: {start_date} hasta: {end_date}\n")
            if title or (start_date and end_date):
                f.write("\n")

            if not data:
                return filepath

            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(
                f,
                fieldnames=fieldnames,
                delimiter=";",
                quoting=csv.QUOTE_MINIMAL,
            )
            writer.writeheader()
            writer.writerows(data)

        return filepath

    @staticmethod
    def export_excel(data: list[dict], filepath: str, start_date: str = "", end_date: str = "", title: str = "") -> str:
        """Write *data* (list of dicts) to an Excel (.xlsx) file using openpyxl.

        Args:
            data:     List of homogeneous dicts (all same keys).
            filepath: Destination path for the Excel file.
            start_date: Start date of the report range.
            end_date:   End date of the report range.
            title:      Optional title for the report.

        Returns:
            The *filepath* on success.
        """
        import openpyxl
        import os
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        row_offset = 1
        if title:
            ws.cell(row=row_offset, column=1, value=title)
            row_offset += 1
        if start_date and end_date:
            ws.cell(row=row_offset, column=1, value=f"Reporte desde: {start_date} hasta: {end_date}")
            row_offset += 1
        if title or (start_date and end_date):
            ws.cell(row=row_offset, column=1, value="")  # spacer row
            row_offset += 1

        if data:
            # Write headers
            headers = list(data[0].keys())
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_offset, column=col_idx, value=h)

            # Write rows
            for row_idx, row in enumerate(data, row_offset + 1):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row[h])

        wb.save(filepath)
        return filepath

    @staticmethod
    def export_pdf(data: list[dict], filepath: str, start_date: str = "", end_date: str = "", title: str = "") -> str:
        """Write *data* (list of dicts) to a PDF (.pdf) file using fpdf2.

        Args:
            data:       List of dicts representing key-value pairs (e.g. Concepto and Monto).
            filepath:   Destination path for the PDF file.
            start_date: Start date of the report range.
            end_date:   End date of the report range.
            title:      Optional title for the report.

        Returns:
            The *filepath* on success.
        """
        from fpdf import FPDF
        import os
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_margins(10, 10, 10)

        # Title
        display_title = title if title else "Reporte"
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, display_title, ln=True, align="L")

        # Period
        if start_date and end_date:
            pdf.set_font("Helvetica", "", 12)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 8, f"Periodo: {start_date} hasta {end_date}", ln=True, align="L")
            pdf.set_text_color(0, 0, 0)
        
        pdf.ln(5)

        if data:
            headers = list(data[0].keys())
            
            # Calculate dynamic column widths based on content
            pdf.set_font("Helvetica", "B", 10)
            col_widths = []
            for header in headers:
                max_w = pdf.get_string_width(str(header)) + 4
                pdf.set_font("Helvetica", "", 10)
                for row in data:
                    val = str(row.get(header, ""))
                    w = pdf.get_string_width(val) + 4
                    if w > max_w:
                        max_w = w
                col_widths.append(max_w)
                pdf.set_font("Helvetica", "B", 10)
                
            # Scale column widths to fit the page (190mm)
            usable_width = 190
            total_w = sum(col_widths)
            if total_w > 0:
                scale = usable_width / total_w
                col_widths = [w * scale for w in col_widths]
            else:
                col_widths = [usable_width / len(headers)] * len(headers)
                
            def truncate_text(text, max_w):
                if pdf.get_string_width(text) <= max_w:
                    return text
                while len(text) > 0 and pdf.get_string_width(text + "...") > max_w:
                    text = text[:-1]
                return text + "..."
            
            # 1. Header
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_fill_color(240, 240, 240)
            for i, header in enumerate(headers):
                ln = True if i == len(headers) - 1 else False
                display_header = truncate_text(str(header), col_widths[i] - 2)
                pdf.cell(col_widths[i], 10, display_header, border=1, ln=ln, align="C", fill=True)

            # 2. Rows
            pdf.set_font("Helvetica", "", 10)
            for row in data:
                # To highlight in Resumen de Ingresos y Egresos specifically
                concept = str(row.get("Concepto", ""))
                
                # Check for negative amounts to highlight in red (Requested to be removed by Gerente)
                # has_negative = any(isinstance(v, str) and "-" in v for v in row.values())
                
                if concept in ["Ganancia Bruta", "Ganancia Neta"]:
                    pdf.set_text_color(31, 111, 58) # Green
                else:
                    pdf.set_text_color(0, 0, 0) # Black
                
                for i, header in enumerate(headers):
                    # Replace em-dash with hyphen to avoid fpdf2 latin-1 encoding errors
                    val = str(row.get(header, "")).replace('\u2014', '-')
                    ln = True if i == len(headers) - 1 else False
                    
                    # Try to right align if it looks like a number
                    align = "R" if (val.startswith("$") or val.replace(".","").replace(",","").replace("-","").isdigit()) else "L"
                    
                    display_val = truncate_text(val, col_widths[i] - 2)
                    pdf.cell(col_widths[i], 10, display_val, border=1, ln=ln, align=align)
                    
                pdf.set_text_color(0, 0, 0) # Reset color

        pdf.output(filepath)
        return filepath
