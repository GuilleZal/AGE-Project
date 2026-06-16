"""Excel import controller — validate, preview, and execute transactional Excel imports.

Handles the full import lifecycle: header validation → row-by-row parsing
and validation → intra-file duplicate detection → transactional upsert.
All valid rows are applied atomically — any DB error rolls back the entire import.
"""

import sqlite3

from pos.model.enums import UnitType
from pos.model.exceptions import POSException
from pos.model.product import Product
from pos.repository.product_repo import ProductRepo
from pos.repository.category_repo import CategoryRepo


# Expected columns in the Excel file (exact match required, in Spanish)
_EXPECTED_HEADERS = ["codigo", "nombre", "categoria", "precio_venta", "precio_costo", "stock", "tipo_unidad"]

# Mapping from Spanish headers to internal field names
_HEADER_MAP = {
    "codigo": "barcode",
    "nombre": "name",
    "categoria": "category_name",
    "precio_venta": "sale_price",
    "precio_costo": "cost_price",
    "stock": "stock",
    "tipo_unidad": "unit_type",
}


class ExcelImportController:
    """Orchestrates Excel product imports with validation and transactional upsert."""

    def __init__(self, db: sqlite3.Connection) -> None:
        self._db = db
        self._product_repo = ProductRepo(db)
        self._category_repo = CategoryRepo(db)
        self._last_result: dict | None = None

    # ---------------------------------------------------------- validate ----

    def validate_excel(self, file_path: str) -> dict:
        """Validate the Excel file structure and each row's data.

        Returns ``{"success": True, "data": {valid_rows, errors}, "error": None}``
        or an error dict if the file cannot be opened or headers are wrong.
        """
        try:
            rows, header_errors = _read_and_validate_headers(file_path)
            if header_errors:
                return {"success": False, "data": None, "error": header_errors[0]}

            valid_rows, row_errors = _validate_all_rows(rows)
            if row_errors:
                return {
                    "success": True,
                    "data": {"valid_rows": valid_rows, "errors": row_errors},
                    "error": None,
                }

            return {
                "success": True,
                "data": {"valid_rows": valid_rows, "errors": []},
                "error": None,
            }
        except FileNotFoundError:
            return {"success": False, "data": None, "error": f"Archivo no encontrado: {file_path}"}
        except Exception as e:
            return {"success": False, "data": None, "error": f"Error al leer archivo: {e}"}

    # ----------------------------------------------------------- preview ----

    def preview_import(self, file_path: str) -> dict:
        """Return the first 10 rows of the Excel file for preview.

        Returns ``{"success": True, "data": {headers, rows, total_rows}, "error": None}``.
        """
        try:
            rows, header_errors = _read_and_validate_headers(file_path)
            if header_errors:
                return {"success": False, "data": None, "error": header_errors[0]}

            total = len(rows)
            preview = rows[:10]

            return {
                "success": True,
                "data": {
                    "headers": _EXPECTED_HEADERS,
                    "rows": preview,
                    "total_rows": total,
                },
                "error": None,
            }
        except FileNotFoundError:
            return {"success": False, "data": None, "error": f"Archivo no encontrado: {file_path}"}
        except Exception as e:
            return {"success": False, "data": None, "error": f"Error al leer archivo: {e}"}

    # ------------------------------------------------------- execute import --

    def execute_import(self, file_path: str) -> dict:
        """Execute the full transactional import with detailed error reporting.

        Steps:
        1. Read and validate headers.
        2. Parse and validate all rows (skip invalid rows, record errors).
        3. Detect and flag intra-file barcode duplicates.
        4. Process each valid row individually with try/except:
           - If successful: upsert product
           - If error: capture error details and continue with next row
        5. Commit all successful operations.

        Returns:
            ``{"success": True, "data": {created, updated, errors, error_details}, "error": None}``
            where ``created`` and ``updated`` are counts, ``errors`` is the count of errors,
            and ``error_details`` is a list of detailed error messages.
        """
        try:
            rows, header_errors = _read_and_validate_headers(file_path)
            if header_errors:
                return {"success": False, "data": None, "error": header_errors[0]}

            valid_rows, row_errors = _validate_all_rows(rows)

            # Intra-file barcode duplicates
            barcode_counts: dict[str, int] = {}
            for i, row in enumerate(valid_rows):
                barcode = str(row["barcode"] or "").strip()
                if barcode:
                    barcode_counts[barcode] = barcode_counts.get(barcode, 0) + 1

            duplicate_barcodes = {b for b, c in barcode_counts.items() if c > 1}
            if duplicate_barcodes:
                # Flag all rows with duplicate barcodes as errors
                final_rows = []
                for i, row in enumerate(valid_rows):
                    barcode = str(row["barcode"] or "").strip()
                    if barcode in duplicate_barcodes:
                        row_errors.append({
                            "row": i + 2,  # +2 for header + 1-indexed
                            "field": "barcode",
                            "value": barcode,
                            "error": "Código de barras duplicado dentro del archivo",
                        })
                    else:
                        final_rows.append(row)
            else:
                final_rows = valid_rows

            if not final_rows:
                return {
                    "success": True,
                    "data": {
                        "created": 0,
                        "updated": 0,
                        "errors": len(row_errors),
                        "error_details": row_errors
                    },
                    "error": None,
                }

            # --- Row-by-row processing with individual error handling ---
            created = 0
            updated = 0
            error_details = list(row_errors)  # Start with validation errors
            
            try:
                self._db.execute("BEGIN")

                for idx, row in enumerate(final_rows):
                    row_num = idx + 2  # +2 for header + 1-indexed
                    try:
                        # Sanitize and resolve category name to category_id
                        category_name = str(row.get("category_name", "") or "").strip()
                        category_id = None
                        if category_name:
                            # Sanitize category name: trim and title case
                            sanitized_name = _sanitize_category_name(category_name)
                            # Try to find existing category
                            existing_cat = self._category_repo.find_by_name(sanitized_name)
                            if existing_cat:
                                category_id = existing_cat.id
                            else:
                                # Create new category
                                new_cat = self._category_repo.create(sanitized_name)
                                category_id = new_cat.id

                        product = Product(
                            barcode=str(row["barcode"]).strip() if row["barcode"] else None,
                            name=str(row["name"]).strip(),
                            category_id=category_id,
                            sale_price=int(row["sale_price"]),
                            cost_price=int(row["cost_price"]),
                            stock=float(row["stock"]),
                            unit_type=str(row["unit_type"]).strip(),
                        )
                        _, action = self._product_repo.upsert_from_import(product)
                        if action == "created":
                            created += 1
                        else:
                            updated += 1
                    except Exception as e:
                        # Capture error for this specific row and continue
                        error_details.append({
                            "row": row_num,
                            "field": "general",
                            "value": None,
                            "error": f"Error al procesar fila: {str(e)}",
                        })

                self._db.execute("COMMIT")
            except Exception:
                self._db.execute("ROLLBACK")
                return {
                    "success": False,
                    "data": None,
                    "error": "Error durante la importación. Se revirtieron todos los cambios.",
                }

            result = {
                "created": created,
                "updated": updated,
                "errors": len(error_details),
                "error_details": error_details,
            }
            self._last_result = result
            return {"success": True, "data": result, "error": None}

        except FileNotFoundError:
            return {"success": False, "data": None, "error": f"Archivo no encontrado: {file_path}"}
        except Exception as e:
            return {"success": False, "data": None, "error": f"Error al importar: {e}"}

    # ------------------------------------------------------- last result ----

    def get_import_result(self) -> dict:
        """Return the result of the last ``execute_import`` call.

        Returns ``{"success": True, "data": result, "error": None}``
        or ``{"success": False, "data": None, "error": "..."}`` if no import has been run.
        """
        if self._last_result is None:
            return {
                "success": False,
                "data": None,
                "error": "No se ha ejecutado ninguna importación",
            }
        return {"success": True, "data": self._last_result, "error": None}


# ----------------------------------------------------------------- helpers ---

def _read_and_validate_headers(file_path: str) -> tuple[list[dict], list[str]]:
    """Open an .xlsx file and validate headers.

    Returns:
        ``(rows, errors)`` — *rows* is a list of dicts keyed by internal field names,
        *errors* is a list of error strings (empty if valid).
    """
    if not file_path.lower().endswith(".xlsx"):
        return [], ["Formato no soportado. Use un archivo .xlsx"]

    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    raw_headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if raw_headers != _EXPECTED_HEADERS:
        return [], [
            f"Encabezados incorrectos. Se esperaba: {', '.join(_EXPECTED_HEADERS)}. "
            f"Encontrado: {', '.join(raw_headers)}"
        ]

    # Read data rows
    rows: list[dict] = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row_cells):
            continue  # skip fully empty rows
        row_dict = {}
        for i, header in enumerate(_EXPECTED_HEADERS):
            value = row_cells[i] if i < len(row_cells) else None
            # Map Spanish header to internal field name
            internal_name = _HEADER_MAP[header]
            row_dict[internal_name] = value
        rows.append(row_dict)

    wb.close()
    return rows, []


def _validate_all_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Validate every row and split into valid / errors.

    Returns:
        ``(valid_rows, error_entries)`` where each error entry is
        ``{"row": n, "field": str, "value": any, "error": str}``.
    """
    valid: list[dict] = []
    errors: list[dict] = []

    for idx, row in enumerate(rows):
        row_num = idx + 2  # 1-indexed, offset for header
        row_errors = _validate_single_row(row, row_num)
        if row_errors:
            errors.extend(row_errors)
        else:
            valid.append(row)

    return valid, errors


def _validate_single_row(row: dict, row_num: int) -> list[dict]:
    """Validate a single data row. Returns a list of error dicts (empty = valid)."""
    errs: list[dict] = []

    # name is required
    name = str(row.get("name", "") or "").strip()
    if not name:
        errs.append({"row": row_num, "field": "nombre", "value": name, "error": "El nombre es obligatorio"})

    # sale_price: must be numeric and ≥ 0
    try:
        sp = float(row.get("sale_price", -1))
        if sp < 0:
            errs.append({"row": row_num, "field": "precio_venta", "value": row.get("sale_price"), "error": "El precio debe ser ≥ 0"})
    except (ValueError, TypeError):
        errs.append({"row": row_num, "field": "precio_venta", "value": row.get("sale_price"), "error": "Precio no numérico"})

    # cost_price: must be numeric and ≥ 0
    try:
        cp = float(row.get("cost_price", -1))
        if cp < 0:
            errs.append({"row": row_num, "field": "precio_costo", "value": row.get("cost_price"), "error": "El costo debe ser ≥ 0"})
    except (ValueError, TypeError):
        errs.append({"row": row_num, "field": "precio_costo", "value": row.get("cost_price"), "error": "Costo no numérico"})

    # stock: must be numeric and ≥ 0
    try:
        st = float(row.get("stock", -1))
        if st < 0:
            errs.append({"row": row_num, "field": "stock", "value": row.get("stock"), "error": "El stock debe ser ≥ 0"})
    except (ValueError, TypeError):
        errs.append({"row": row_num, "field": "stock", "value": row.get("stock"), "error": "Stock no numérico"})

    # unit_type: must be one of the valid values (normalize common abbreviations)
    _UNIT_ALIASES = {"u": "unit", "uni": "unit", "kg": "weight_kg", "p": "pack", "paq": "pack"}
    ut = str(row.get("unit_type", "") or "").strip().lower()
    ut = _UNIT_ALIASES.get(ut, ut)
    row["unit_type"] = ut
    valid_units = {"unit", "weight_kg", "pack"}
    if ut not in valid_units:
        errs.append({"row": row_num, "field": "tipo_unidad", "value": ut, "error": f"Tipo inválido. Use: unit, weight_kg, pack"})

    return errs


def _sanitize_category_name(name: str) -> str:
    """Sanitize a category name by trimming whitespace and applying title case.
    
    Examples:
        " bebidas " -> "Bebidas"
        "BEBIDAS" -> "Bebidas"
        "bebidas" -> "Bebidas"
        "  ALIMENTOS  " -> "Alimentos"
    
    Args:
        name: Raw category name from Excel.
    
    Returns:
        Sanitized category name with title case.
    """
    # Trim whitespace
    sanitized = name.strip()
    # Apply title case (first letter of each word capitalized)
    sanitized = sanitized.title()
    return sanitized
