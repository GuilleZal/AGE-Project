import pytest
import sqlite3
from pos.controller.cash_register_controller import CashRegisterController
from pos.model.enums import MovementType
from pos.repository.cash_movement_repo import CashMovementRepo


@pytest.fixture
def cash_ctrl(db: sqlite3.Connection) -> CashRegisterController:
    return CashRegisterController(db)


def test_format_movements_with_discount(db: sqlite3.Connection, cash_ctrl: CashRegisterController):
    # 1. Open register
    cash_ctrl.open_register(5000)
    active_reg = cash_ctrl._register_repo.find_active()
    assert active_reg is not None

    # 2. Insert a sale with discount
    # total = 900, discount = 100 (subtotal = 1000, discount_pct = 10.0%)
    db.execute(
        """INSERT INTO sales (total, discount, surcharge, payment_method, cash_register_id)
           VALUES (?, ?, ?, ?, ?)""",
        (900, 100, 0, "cash", active_reg.id)
    )
    sale_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    # 3. Create cash movement
    mov_repo = CashMovementRepo(db)
    movement = mov_repo.create(
        register_id=active_reg.id,
        type_=MovementType.SALE_CASH,
        amount=900,
        description=f"Venta #{sale_id}"
    )

    # 4. Format movements
    formatted = cash_ctrl._format_movements_for_display([movement])
    assert len(formatted) == 1
    assert formatted[0]["discount_pct"] == 10.0
    assert formatted[0]["description"] == "Venta #1"


def test_sale_view_cashier_payment_method_labels(session_root):
    from pos.view.sale_view import SaleView
    
    # 1. Non-cashier view (default or admin)
    view_admin = SaleView(session_root, role="admin")
    session_root.update()
    
    # Check if "Efectivo" and "Transferencia" are present
    labels_admin = []
    def collect_labels(widget, target_list):
        import customtkinter as ctk
        if isinstance(widget, ctk.CTkLabel):
            target_list.append(widget.cget("text"))
        for child in widget.winfo_children():
            collect_labels(child, target_list)
            
    collect_labels(view_admin, labels_admin)
    assert "Efectivo" in labels_admin
    assert "Transferencia" in labels_admin
    view_admin.destroy()
    
    # 2. Cashier view ("cajero")
    view_cajero = SaleView(session_root, role="cajero")
    session_root.update()
    
    labels_cajero = []
    collect_labels(view_cajero, labels_cajero)
    assert "Eritrea" not in labels_cajero  # arbitrary check to verify collect
    assert "Efectivo" in labels_cajero
    assert "Transferencia" in labels_cajero
    assert "Qr" in labels_cajero
    view_cajero.destroy()


def test_report_controller_export_to_excel(tmp_path):
    import os
    import openpyxl
    from pos.controller.report_controller import ReportController
    
    # We don't need a real db for this test since export_to_excel doesn't touch sqlite
    ctrl = ReportController(None)
    
    data = [
        {"Nro": 1, "Producto": "Coca Cola 1.5L", "Cantidad": 5, "Monto Total": "$5000"},
        {"Nro": 2, "Producto": "Pepsi 1.5L", "Cantidad": 3, "Monto Total": "$3000"},
    ]
    
    dest_file = tmp_path / "reporte.xlsx"
    result = ctrl.export_to_excel(data, str(dest_file), "2026-07-27", "2026-07-28")
    
    assert result["success"] is True
    assert os.path.exists(dest_file)
    
    # Verify contents with openpyxl
    wb = openpyxl.load_workbook(str(dest_file))
    ws = wb.active
    assert ws.title == "Reporte"
    
    # Check date range header
    assert ws.cell(row=1, column=1).value == "Reporte desde: 2026-07-27 hasta: 2026-07-28"
    assert ws.cell(row=2, column=1).value is None
    
    # Check headers
    headers = [cell.value for cell in ws[3]]
    assert headers == ["Nro", "Producto", "Cantidad", "Monto Total"]
    
    # Check data row 1
    row1 = [cell.value for cell in ws[4]]
    assert row1 == [1, "Coca Cola 1.5L", 5, "$5000"]
    
    # Check data row 2
    row2 = [cell.value for cell in ws[5]]
    assert row2 == [2, "Pepsi 1.5L", 3, "$3000"]


def test_report_controller_export_to_csv_with_dates(tmp_path):
    import os
    from pos.controller.report_controller import ReportController
    
    ctrl = ReportController(None)
    data = [
        {"Nro": 1, "Producto": "Coca Cola 1.5L", "Cantidad": 5, "Monto Total": "$5000"},
    ]
    
    dest_file = tmp_path / "reporte.csv"
    result = ctrl.export_to_csv(data, str(dest_file), "2026-07-27", "2026-07-28")
    
    assert result["success"] is True
    assert os.path.exists(dest_file)
    
    content = dest_file.read_text(encoding="utf-8-sig")
    assert "Reporte desde: 2026-07-27 hasta: 2026-07-28" in content
    assert "Nro;Producto;Cantidad;Monto Total" in content
    assert "1;Coca Cola 1.5L;5;$5000" in content


def test_report_summary_dialog_hides_csv_export_for_gerente(session_root):
    import customtkinter as ctk
    from pos.view.widgets.report_summary_dialog import ReportSummaryDialog
    
    root = session_root
    report_data = {
        "expenses": {
            "purchases": 100,
            "shrinkage": 10,
            "operating_expenses": 20,
            "net_profit": 50,
        }
    }
    
    # Dialog for gerente
    dialog_gerente = ReportSummaryDialog(root, report_data, role="gerente")
    
    # Find all CTkButtons in dialog_gerente
    def find_buttons(widget):
        buttons = []
        if isinstance(widget, ctk.CTkButton):
            buttons.append(widget)
        for child in widget.winfo_children():
            buttons.extend(find_buttons(child))
        return buttons
        
    btns_gerente = find_buttons(dialog_gerente)
    # Gerente should see "Cerrar" and "Exportar (PDF)", but no "Exportar (Excel)" or "Exportar Resumen (CSV)"
    button_texts_gerente = [btn.cget("text") for btn in btns_gerente]
    assert "Cerrar" in button_texts_gerente
    assert "Exportar (PDF)" in button_texts_gerente
    assert "Exportar (Excel)" not in button_texts_gerente
    assert "Exportar Resumen (CSV)" not in button_texts_gerente
    
    # Dialog for admin
    dialog_admin = ReportSummaryDialog(root, report_data, role="admin")
    btns_admin = find_buttons(dialog_admin)
    button_texts_admin = [btn.cget("text") for btn in btns_admin]
    assert "Cerrar" in button_texts_admin
    assert "Exportar Resumen (CSV)" in button_texts_admin
    
    dialog_gerente.destroy()
    dialog_admin.destroy()


def test_report_view_uses_cached_dates_on_export(session_root, mocker):
    import customtkinter as ctk
    from pos.view.report_view import ReportView
    
    # Mock messagebox to prevent blocking dialogs
    mocker.patch("pos.view.report_view.messagebox.showinfo")
    mocker.patch("pos.view.report_view.messagebox.showerror")
    
    root = session_root
    view = ReportView(root, role="gerente")
    
    # Mock date range to return a changed period
    view._get_date_range = lambda: ("2026-08-01 00:00:00", "2026-08-02 23:59:59")
    
    # Set the cached (actually generated) dates
    view._generated_start_date = "2026-07-01"
    view._generated_end_date = "2026-07-15"
    
    # Mock controller and its export function
    mock_controller = mocker.Mock()
    mock_controller.export_to_excel.return_value = {"success": True, "data": "test.xlsx", "error": None}
    view.set_controller(mock_controller)
    
    # Trigger export
    view._execute_export([{"col": "val"}], "test.xlsx", "Title")
    
    # Verify that export_to_excel was called with the CACHED dates, not the current calendar dates
    mock_controller.export_to_excel.assert_called_once_with(
        [{"col": "val"}],
        "test.xlsx",
        "2026-07-01",
        "2026-07-15"
    )
    
    view.destroy()


def test_report_summary_dialog_pdf_export_call(session_root, mocker):
    from pos.view.widgets.report_summary_dialog import ReportSummaryDialog
    
    root = session_root
    report_data = {
        "period": {"start": "2026-07-01", "end": "2026-07-15"},
        "sales": {"total": 500},
        "expenses": {"purchases": 100, "shrinkage": 10}
    }
    
    # Mock filedialog.asksaveasfilename
    mocker.patch("pos.view.widgets.report_summary_dialog.filedialog.asksaveasfilename", return_value="test.pdf")
    
    # Mock messagebox.showinfo
    mock_showinfo = mocker.patch("pos.view.widgets.report_summary_dialog.messagebox.showinfo")
    
    # Mock controller and export_to_pdf
    mock_controller = mocker.Mock()
    mock_controller.export_to_pdf.return_value = {"success": True, "data": "test.pdf", "error": None}
    
    root._controller = mock_controller
    
    dialog = ReportSummaryDialog(root, report_data, role="gerente")
    dialog._export_summary_pdf()
    
    # Assert export_to_pdf was called
    mock_controller.export_to_pdf.assert_called_once()
    mock_showinfo.assert_called_once()
    
    dialog.destroy()


def test_report_view_table_export_buttons_visibility(session_root):
    from pos.view.report_view import ReportView
    
    root = session_root
    
    # Gerente view
    view_gerente = ReportView(root, role="gerente")
    session_root.update()
    
    # Gerente should have CSV button gridded but NOT Excel button
    assert view_gerente._export_table_btn.grid_info()
    assert not view_gerente._export_excel_btn.grid_info()
    
    view_gerente.destroy()
    
    # Admin view
    view_admin = ReportView(root, role="admin")
    session_root.update()
    
    # Admin should have BOTH CSV and Excel buttons gridded
    assert view_admin._export_table_btn.grid_info()
    assert view_admin._export_excel_btn.grid_info()
    
    view_admin.destroy()
